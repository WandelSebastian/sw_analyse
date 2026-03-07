import openpyxl, json, re

wb = openpyxl.load_workbook('Vorlage_Krafttraining_Alle_Level.xlsx', data_only=True)

# --- 1. Parse exercise sheets ---
sheet_map = {}
for sn in wb.sheetnames:
    m = re.match(r'(Unterk.rper|Oberk.rper)(\w+)', sn.strip())
    if m:
        bp = 'lowerBody' if 'nter' in m.group(1) else 'upperBody'
        level = m.group(2)
        sheet_map[sn] = (level, bp)

BLOCK_PATTERNS = [
    (r'Explosiv\s*Block', 'explosiv'),
    (r'Kraft\s*Block\s*\(UKK\)', 'strengthA'),
    (r'Kraft\s*Block\s*\(OKK\)', 'strengthA'),
    (r'Kraft\s*Block\s*\(Ukp\)', 'strengthB'),
    (r'Kraft\s*Block\s*\(Okp\)', 'strengthB'),
    (r'Halte.bungen', 'isometrics'),
]

def identify_block(text):
    if not text:
        return None
    for pattern, block_id in BLOCK_PATTERNS:
        if re.search(pattern, text, re.IGNORECASE):
            return block_id
    return None

def level_sort_key(lvl):
    if lvl.isdigit():
        return (1, int(lvl))
    return (0, ord(lvl))

exercises_data = {"levels": {}}

for sn, (level, bp) in sorted(sheet_map.items(), key=lambda x: x[1]):
    ws = wb[sn]
    if level not in exercises_data["levels"]:
        exercises_data["levels"][level] = {}
    if bp not in exercises_data["levels"][level]:
        exercises_data["levels"][level][bp] = {}

    current_block = None
    exercises = []

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]

        # Check if this row is a block header
        col_b = str(vals[1]) if vals[1] else ''
        block_id = identify_block(col_b)

        if block_id:
            if current_block and exercises:
                if current_block not in exercises_data["levels"][level][bp]:
                    exercises_data["levels"][level][bp][current_block] = []
                exercises_data["levels"][level][bp][current_block].extend(exercises)
            current_block = block_id
            exercises = []
            continue

        if current_block is None:
            continue

        # Check if this is an exercise row (col B has a number)
        try:
            ex_order = int(vals[1])
        except (ValueError, TypeError):
            continue

        ex_name = str(vals[2]).strip() if vals[2] else ''
        if not ex_name:
            continue

        # Extract tempo/RPE/SxR from first block columns
        tempo = str(vals[6]).strip() if vals[6] else ''
        rpe_or_tempo = str(vals[7]).strip() if vals[7] else ''
        sxr = str(vals[8]).strip() if vals[8] else ''
        weight = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ''

        # Build exercise ID
        bp_short = 'lb' if bp == 'lowerBody' else 'ub'
        block_short = {'explosiv': 'e', 'strengthA': 's', 'strengthB': 'p', 'isometrics': 'i'}
        bs = block_short.get(current_block, 'x')
        ex_id = f"ex-{level.lower()}-{bp_short}-{bs}{ex_order}"

        ex = {"id": ex_id, "order": ex_order, "name": ex_name}

        if current_block == 'explosiv':
            ex["tempo"] = tempo if tempo else "EXPL"
            ex["defaultSxR"] = sxr
            if weight:
                ex["defaultWeight"] = weight
        else:
            if tempo:
                ex["tempo"] = tempo
            if rpe_or_tempo:
                ex["defaultRPE"] = rpe_or_tempo
            if sxr:
                ex["defaultSxR"] = sxr
            if weight:
                ex["defaultWeight"] = weight

        exercises.append(ex)

    # Save last block
    if current_block and exercises:
        if current_block not in exercises_data["levels"][level][bp]:
            exercises_data["levels"][level][bp][current_block] = []
        exercises_data["levels"][level][bp][current_block].extend(exercises)

# Sort levels
sorted_levels = dict(sorted(exercises_data["levels"].items(), key=lambda x: level_sort_key(x[0])))
exercises_data["levels"] = sorted_levels

# --- 2. Parse progressions ---
progressions_data = {"lowerBody": [], "upperBody": []}

for sn in wb.sheetnames:
    if 'Progressionen' not in sn:
        continue
    ws = wb[sn]
    if 'BH' in sn:
        bp = 'lowerBody'  # BH progressions go to lowerBody
    elif 'nter' in sn:
        bp = 'lowerBody'
    else:
        bp = 'upperBody'

    headers = [str(c.value).strip() if c.value else '' for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]
        if not vals[0]:
            continue
        level_name = str(vals[0]).strip()
        for col_idx in range(1, min(len(headers), len(vals))):
            if not headers[col_idx]:
                continue
            cat = headers[col_idx]
            ex_name = str(vals[col_idx]).strip() if vals[col_idx] else ''
            if not ex_name:
                continue
            # Find or create progression
            found = None
            for p in progressions_data[bp]:
                if p["name"] == cat:
                    found = p
                    break
            if not found:
                found = {"name": cat, "levels": []}
                progressions_data[bp].append(found)
            found["levels"].append({"level": level_name, "exercise": ex_name})

# --- 3. Parse warmup (BH1/BH2) from exercise sheets ---
warmup_data = {"categories": []}
bh_exercises = {}

for sn, (level, bp) in sheet_map.items():
    if bp != 'lowerBody':
        continue
    ws = wb[sn]
    rows_list = list(ws.iter_rows(min_row=1, max_row=20, values_only=False))
    for i, row in enumerate(rows_list):
        vals = [c.value for c in row]
        col_b = str(vals[1]) if vals[1] else ''
        if 'Bewegungshygiene 1' in col_b or 'Movement Patterns 1' in col_b:
            if i + 1 < len(rows_list):
                next_vals = [c.value for c in rows_list[i + 1]]
                exercises_list = [str(v).strip() for v in next_vals if v and str(v).strip()]
                if 'BH1' not in bh_exercises:
                    bh_exercises['BH1'] = {}
                bh_exercises['BH1'][level] = exercises_list
        elif 'Bewegungshygiene 2' in col_b or 'Movement Patterns 2' in col_b:
            if i + 1 < len(rows_list):
                next_vals = [c.value for c in rows_list[i + 1]]
                exercises_list = [str(v).strip() for v in next_vals if v and str(v).strip()]
                if 'BH2' not in bh_exercises:
                    bh_exercises['BH2'] = {}
                bh_exercises['BH2'][level] = exercises_list

for bh_key in sorted(bh_exercises.keys()):
    cat = {"name": bh_key, "levels": []}
    for lvl in sorted(bh_exercises[bh_key].keys(), key=level_sort_key):
        cat["levels"].append({"level": lvl, "exercises": bh_exercises[bh_key][lvl]})
    warmup_data["categories"].append(cat)

# --- Write output ---
with open('frontend/public/data/exercises.json', 'w', encoding='utf-8') as f:
    json.dump(exercises_data, f, ensure_ascii=False, indent=2)

with open('frontend/public/data/progressions.json', 'w', encoding='utf-8') as f:
    json.dump(progressions_data, f, ensure_ascii=False, indent=2)

with open('frontend/public/data/warmup.json', 'w', encoding='utf-8') as f:
    json.dump(warmup_data, f, ensure_ascii=False, indent=2)

# Summary
total_ex = 0
for lvl in exercises_data["levels"]:
    for bp_key in exercises_data["levels"][lvl]:
        for block in exercises_data["levels"][lvl][bp_key]:
            n = len(exercises_data["levels"][lvl][bp_key][block])
            total_ex += n

print(f"Done!")
print(f"Levels: {list(exercises_data['levels'].keys())}")
print(f"Total exercises: {total_ex}")
for lvl in exercises_data["levels"]:
    blocks = {}
    for bp_key in exercises_data["levels"][lvl]:
        for block in exercises_data["levels"][lvl][bp_key]:
            blocks[f"{bp_key}/{block}"] = len(exercises_data["levels"][lvl][bp_key][block])
    print(f"  Level {lvl}: {blocks}")
print(f"Progressions LB: {len(progressions_data['lowerBody'])}, UB: {len(progressions_data['upperBody'])}")
print(f"Warmup categories: {len(warmup_data['categories'])}")
