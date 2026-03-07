"""
Import exercises from Vorlage_Krafttraining_Alle_Level.xlsx into the backend API.

Creates deduplicated exercise master records and level assignments.
"""
import openpyxl, json, re, requests, sys

API = "http://localhost:8082/api/v1"
EXCEL = "Vorlage_Krafttraining_Alle_Level.xlsx"

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# --- Map sheet names ---
sheet_map = {}
for sn in wb.sheetnames:
    m = re.match(r'(Unterk.rper|Oberk.rper)(\w+)', sn.strip())
    if m:
        bp = 'lowerBody' if 'nter' in m.group(1) else 'upperBody'
        level = m.group(2)
        sheet_map[sn] = (level, bp)

BODY_PREFIX = {'lowerBody': 'uk', 'upperBody': 'ok', 'core': 'core', 'fullBody': 'uk'}

# Pattern -> (block_suffix, category)
BLOCK_PATTERNS = [
    (r'Explosiv\s*Block', 'ex', 'EX'),
    (r'Kraft\s*Block\s*\(UKK\)', 'k', 'K'),
    (r'Kraft\s*Block\s*\(OKK\)', 'k', 'K'),
    (r'Kraft\s*Block\s*\(Ukp\)', 'p', 'P'),
    (r'Kraft\s*Block\s*\(Okp\)', 'p', 'P'),
    (r'Halte.bungen', 'iso', 'ISO'),
]

def identify_block(text, body_part):
    """Returns (block_id, category) where block_id = bodyPrefix + suffix."""
    if not text:
        return None, None
    prefix = BODY_PREFIX.get(body_part, 'uk')
    for pattern, suffix, category in BLOCK_PATTERNS:
        if re.search(pattern, text, re.IGNORECASE):
            return prefix + suffix, category
    return None, None

def normalize_name(name):
    """Normalize exercise name for deduplication."""
    return re.sub(r'\s+', ' ', name.strip().upper())

# --- Pass 1: Collect all exercises and deduplicate ---
# Key: normalized name -> exercise master record
exercise_master = {}  # norm_name -> {id, name, bodyRegion, category, ...}
# All level assignments
level_assignments = []

def get_or_create_exercise(name, body_region, category):
    norm = normalize_name(name)
    if norm in exercise_master:
        ex = exercise_master[norm]
        # If seen in both body regions, mark as fullBody
        if ex["bodyRegion"] != body_region:
            if ex["bodyRegion"] not in ("fullBody",):
                ex["bodyRegion"] = "fullBody"
        return ex["id"]

    # Create slug-style ID
    slug = re.sub(r'[^a-z0-9]+', '-', name.strip().lower()).strip('-')
    # Ensure unique
    base_slug = slug
    counter = 2
    while any(e["id"] == slug for e in exercise_master.values()):
        slug = f"{base_slug}-{counter}"
        counter += 1

    exercise_master[norm] = {
        "id": slug,
        "name": name.strip(),
        "bodyRegion": body_region,
        "category": category,
        "tags": [],
        "equipment": [],
        "description": "",
    }
    return slug

# --- Parse all exercise sheets ---
for sn, (level, bp) in sorted(sheet_map.items(), key=lambda x: x[1]):
    ws = wb[sn]
    current_block = None
    current_category = None
    # Column layout detected from block header row
    col_layout = None  # 'old' (A-F) or 'new' (1-10)
    header_rpe = ''    # RPE extracted from header like "SxR/RPE 8-9"

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [c.value for c in row]
        col_b = str(vals[1]) if vals[1] else ''
        block_id, cat = identify_block(col_b, bp)

        if block_id:
            current_block = block_id
            current_category = cat
            # Detect column layout from header row
            h6 = str(vals[6]).strip().upper() if vals[6] else ''
            h7 = str(vals[7]).strip().upper() if vals[7] else ''
            h8 = str(vals[8]).strip() if vals[8] else ''
            if h7 == 'RPE':
                col_layout = 'old'  # Levels A-F: col6=TEMPO, col7=RPE, col8=SxR
                header_rpe = ''
            elif h7 == 'TEMPO':
                col_layout = 'new'  # Levels 1-10: col6=GEWICHT, col7=TEMPO, col8=SxR/RPE
                # Extract RPE from header like "SxR/RPE 8-9"
                m = re.search(r'RPE\s*([\d\-]+)', h8, re.IGNORECASE)
                header_rpe = m.group(1) if m else ''
            else:
                col_layout = 'old'
                header_rpe = ''
            continue

        if current_block is None:
            continue

        try:
            ex_order = int(vals[1])
        except (ValueError, TypeError):
            continue

        ex_name = str(vals[2]).strip() if vals[2] else ''
        if not ex_name:
            continue

        # Get or create master exercise
        exercise_id = get_or_create_exercise(ex_name, bp, current_category)

        # Extract training params based on detected column layout
        if col_layout == 'new':
            # Levels 1-10: col6=GEWICHT, col7=TEMPO, col8=SxR, col9=GEWICHT
            weight = str(vals[6]).strip() if vals[6] else ''
            tempo = str(vals[7]).strip() if vals[7] else ''
            sxr = str(vals[8]).strip() if vals[8] else ''
            rpe = header_rpe  # RPE from block header
        else:
            # Levels A-F: col6=TEMPO, col7=RPE, col8=SxR, col9=GEWICHT
            tempo = str(vals[6]).strip() if vals[6] else ''
            rpe = str(vals[7]).strip() if vals[7] else ''
            sxr = str(vals[8]).strip() if vals[8] else ''
            weight = str(vals[9]).strip() if len(vals) > 9 and vals[9] else ''

        # For explosiv blocks, default tempo
        if current_block == 'explosiv' and not tempo:
            tempo = 'EXPL'

        level_assignments.append({
            "exerciseId": exercise_id,
            "level": level,
            "block": current_block,
            "order": ex_order,
            "defaultTempo": tempo,
            "defaultRPE": rpe,
            "defaultSxR": sxr,
            "defaultWeight": weight,
        })

# --- Parse BH (Bewegungshygiene) exercises from warmup rows ---
for sn, (level, bp) in sheet_map.items():
    if bp != 'lowerBody':
        continue
    ws = wb[sn]
    rows_list = list(ws.iter_rows(min_row=1, max_row=20, values_only=False))
    for i, row in enumerate(rows_list):
        vals = [c.value for c in row]
        col_b = str(vals[1]) if vals[1] else ''
        bh_type = None
        if 'Bewegungshygiene 1' in col_b or 'Movement Patterns 1' in col_b:
            bh_type = 'BH'
        elif 'Bewegungshygiene 2' in col_b or 'Movement Patterns 2' in col_b:
            bh_type = 'BH'

        if bh_type and i + 1 < len(rows_list):
            next_vals = [c.value for c in rows_list[i + 1]]
            for v in next_vals:
                if v and str(v).strip():
                    name = str(v).strip()
                    get_or_create_exercise(name, 'lowerBody', 'BH')

print(f"Unique exercises: {len(exercise_master)}")
print(f"Level assignments: {len(level_assignments)}")

# --- Pass 2: Push to API ---
print("\nCreating exercises...")
created = 0
for norm, ex in exercise_master.items():
    r = requests.post(f"{API}/exercises", json=ex)
    if r.status_code in (200, 201):
        created += 1
    else:
        print(f"  WARN: {ex['name']}: {r.status_code} {r.text[:80]}")
print(f"  Created: {created}/{len(exercise_master)}")

print("\nCreating level assignments...")
assigned = 0
for la in level_assignments:
    r = requests.post(f"{API}/level-exercises", json=la)
    if r.status_code in (200, 201):
        assigned += 1
    else:
        print(f"  WARN: {la['exerciseId']} level {la['level']}: {r.status_code} {r.text[:80]}")
print(f"  Created: {assigned}/{len(level_assignments)}")

print("\nDone!")
